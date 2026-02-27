"""
Reports routes - PDF and Excel export.
"""

from datetime import datetime
from io import BytesIO
from flask import Blueprint, send_file, flash, redirect, url_for, request
from flask_login import login_required, current_user
from sqlalchemy import func, extract
from app import db
from app.models.expense import Expense
from app.models.income import Income
from app.models.category import Category

reports_bp = Blueprint('reports', __name__)

@reports_bp.route('/pdf')
@login_required
def download_pdf():
    """Generate and download monthly expense report as PDF."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
    except ImportError:
        flash('ReportLab not installed. Run: pip install reportlab', 'danger')
        return redirect(url_for('main.dashboard'))
    
    month = request.args.get('month', datetime.now().month, type=int)
    year = request.args.get('year', datetime.now().year, type=int)
    
    # Fetch data
    expenses = Expense.query.filter_by(user_id=current_user.user_id).filter(
        extract('month', Expense.expense_date) == month,
        extract('year', Expense.expense_date) == year
    ).join(Category).order_by(Expense.expense_date).all()
    
    income_total = db.session.query(func.sum(Income.amount)).filter(
        Income.user_id == current_user.user_id,
        extract('month', Income.income_date) == month,
        extract('year', Income.income_date) == year
    ).scalar() or 0
    expense_total = sum(e.amount for e in expenses)
    
    # Create PDF in memory
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=72, leftMargin=72)
    styles = getSampleStyleSheet()
    
    elements = []
    title = Paragraph(
        f"<b>Expense Report - {datetime(year, month, 1).strftime('%B %Y')}</b>",
        ParagraphStyle(name='Title', fontSize=16, spaceAfter=20)
    )
    elements.append(title)
    elements.append(Spacer(1, 12))
    
    # Summary
    summary_data = [
        ['Total Income', f'₹{float(income_total):,.2f}'],
        ['Total Expense', f'₹{expense_total:,.2f}'],
        ['Savings', f'₹{float(income_total) - expense_total:,.2f}']
    ]
    t1 = Table(summary_data)
    t1.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(t1)
    elements.append(Spacer(1, 20))
    
    # Expense details
    elements.append(Paragraph("<b>Expense Details</b>", styles['Heading2']))
    elements.append(Spacer(1, 8))
    
    exp_data = [['Date', 'Category', 'Amount', 'Description']]
    for e in expenses:
        exp_data.append([
            e.expense_date.strftime('%Y-%m-%d'),
            e.category.category_name,
            f'₹{e.amount:,.2f}',
            (e.description or '')[:50]
        ])
    
    if len(exp_data) > 1:
        t2 = Table(exp_data, colWidths=[1.2*inch, 1.2*inch, 1*inch, 2.6*inch])
        t2.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        elements.append(t2)
    else:
        elements.append(Paragraph("No expenses for this month.", styles['Normal']))
    
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"expense_report_{year}_{month:02d}.pdf"
    return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=filename)

@reports_bp.route('/excel')
@login_required
def download_excel():
    """Export expenses to Excel file."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        flash('openpyxl not installed. Run: pip install openpyxl', 'danger')
        return redirect(url_for('main.dashboard'))
    
    month = request.args.get('month', datetime.now().month, type=int)
    year = request.args.get('year', datetime.now().year, type=int)
    
    expenses = Expense.query.filter_by(user_id=current_user.user_id).filter(
        extract('month', Expense.expense_date) == month,
        extract('year', Expense.expense_date) == year
    ).join(Category).order_by(Expense.expense_date).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Expenses {month}-{year}"
    
    headers = ['Date', 'Category', 'Amount', 'Description']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
    
    for row, exp in enumerate(expenses, 2):
        ws.cell(row=row, column=1, value=exp.expense_date.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=2, value=exp.category.category_name)
        ws.cell(row=row, column=3, value=exp.amount)
        ws.cell(row=row, column=4, value=exp.description or '')
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"expenses_{year}_{month:02d}.xlsx"
    return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)
